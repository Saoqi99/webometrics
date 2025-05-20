import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
import time
import random
from openpyxl import load_workbook


def extract_names_from_excel(file_path, sheet_name='Sheet6', name_column='Nama2'):
    """
    Mengekstrak nama dari file author79.xlsx sheet ALL
    """
    try:
        # Menggunakan openpyxl untuk handle file xlsx
        wb = load_workbook(filename=file_path)
        sheet = wb[sheet_name]

        # Cari kolom yang berisi nama
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        col_idx = header_row.index(name_column) if name_column in header_row else 0

        names = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[col_idx]:
                names.append(str(row[col_idx]).strip())

        return list(set(names))  # Hapus duplikat
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []


def search_sinta(name, affiliation="UIN Sunan Kalijaga"):
    """
    Mencari peneliti di Sinta (Science and Technology Index) berdasarkan nama dan afiliasi
    """
    base_url = "https://sinta.kemdikbud.go.id/authors"
    params = {
        "search": name,
        "view": "list"
    }

    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        # Random delay sebelum request
        time.sleep(random.uniform(1, 2))

        response = requests.get(base_url, params=params, headers=headers, timeout=30)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        results = []

        for author in soup.select('.author-list-item'):
            author_name_elem = author.select_one('.author-name a')
            if not author_name_elem:
                continue

            author_name = author_name_elem.text.strip()
            author_url = author_name_elem['href']
            author_affiliation = author.select_one('.affil-name').text.strip() if author.select_one(
                '.affil-name') else ""

            # Cek variasi penulisan UIN Sunan Kalijaga
            if any(term.lower() in author_affiliation.lower() for term in
                   ["UIN Sunan Kalijaga", "UIN Suka", "Sunan Kalijaga", "UIN Yogyakarta"]):
                # Dapatkan ID Sinta dari URL
                sinta_id = re.search(r'/authors/(\d+)', author_url)
                if sinta_id:
                    sinta_id = sinta_id.group(1)
                else:
                    sinta_id = ""

                results.append({
                    'original_name': name,
                    'sinta_name': author_name,
                    'sinta_id': sinta_id,
                    'url': author_url,
                    'affiliation': author_affiliation,
                    'scopus_id': author.get('data-scopus-id', ''),
                    'scholar_id': author.get('data-scholar-id', '')
                })

        return results
    except Exception as e:
        print(f"Error searching Sinta for {name}: {e}")
        return []


def get_publications(sinta_url):
    """
    Mendapatkan total publikasi dari tahun 2022-2025
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        # Random delay sebelum request
        time.sleep(random.uniform(1, 2))

        response = requests.get(sinta_url, headers=headers, timeout=30)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        # Cari total publikasi per tahun
        publications = {}
        year_elements = soup.select('.ar-list-item')

        for year_element in year_elements:
            year_text = year_element.select_one('.ar-year').text.strip()
            try:
                year = int(year_text)
                if 2022 <= year <= 2025:
                    count = year_element.select_one('.ar-pub').text.strip()
                    publications[str(year)] = count
            except ValueError:
                continue

        # Hitung total publikasi 2022-2025
        total = sum(int(p) for p in publications.values())

        return {
            'publications_per_year': publications,
            'total_publications': total
        }
    except Exception as e:
        print(f"Error getting publications for {sinta_url}: {e}")
        return {
            'publications_per_year': {},
            'total_publications': 0
        }


def process_names(names, output_file='sinta_results.xlsx'):
    """
    Memproses daftar nama dan menyimpan hasilnya
    """
    results = []
    total_processed = 0
    batch_size = 10  # Jumlah request sebelum istirahat panjang
    batch_count = 0

    for i, name in enumerate(names):
        print(f"Processing {i + 1}/{len(names)}: {name}")
        sinta_accounts = search_sinta(name)

        if not sinta_accounts:
            results.append({
                'Nama Asli': name,
                'Nama di Sinta': '-',
                'ID Sinta': '-',
                'Afiliasi': '-',
                'URL Sinta': '-',
                'Publikasi 2022': '-',
                'Publikasi 2023': '-',
                'Publikasi 2024': '-',
                'Publikasi 2025': '-',
                'Total Publikasi (2022-2025)': '-',
                'Scopus ID': '-',
                'Scholar ID': '-',
                'Status': 'Tidak ditemukan'
            })
        else:
            for account in sinta_accounts:
                publications = get_publications(account['url'])

                results.append({
                    'Nama Asli': account['original_name'],
                    'Nama di Sinta': account['sinta_name'],
                    'ID Sinta': account['sinta_id'],
                    'Afiliasi': account['affiliation'],
                    'URL Sinta': account['url'],
                    'Publikasi 2022': publications['publications_per_year'].get('2022', 0),
                    'Publikasi 2023': publications['publications_per_year'].get('2023', 0),
                    'Publikasi 2024': publications['publications_per_year'].get('2024', 0),
                    'Publikasi 2025': publications['publications_per_year'].get('2025', 0),
                    'Total Publikasi (2022-2025)': publications['total_publications'],
                    'Scopus ID': account['scopus_id'],
                    'Scholar ID': account['scholar_id'],
                    'Status': 'Ditemukan'
                })

        total_processed += 1
        batch_count += 1

        # Istirahat panjang setiap batch_size request
        if batch_count >= batch_size:
            long_delay = random.uniform(1, 30)  # Istirahat 1-30 detik
            print(f"\n--- Menjalankan istirahat panjang selama {long_delay:.1f} detik ---\n")
            time.sleep(long_delay)
            batch_count = 0

    # Simpan hasil ke Excel
    df = pd.DataFrame(results)

    # Urutkan kolom
    columns_order = [
        'Nama Asli', 'Nama di Sinta', 'ID Sinta', 'Afiliasi', 'URL Sinta',
        'Publikasi 2022', 'Publikasi 2023', 'Publikasi 2024', 'Publikasi 2025',
        'Total Publikasi (2022-2025)', 'Scopus ID', 'Scholar ID', 'Status'
    ]
    df = df[columns_order]

    df.to_excel(output_file, index=False)
    print(f"\nHasil disimpan ke {output_file}")
    print(f"Total diproses: {len(names)} nama")
    print(f"Total ditemukan: {len(df[df['Status'] == 'Ditemukan'])} akun Sinta")


def main():
    # File input dan output
    input_file = "author79.xlsx"
    output_file = "hasil_publikasi_sinta.xlsx"

    print("Memulai proses ekstraksi nama dari Excel...")
    names = extract_names_from_excel(input_file, sheet_name='Sheet6')

    if not names:
        print("Tidak ditemukan nama di file Excel.")
        return

    print(f"\nDitemukan {len(names)} nama unik untuk diproses.")
    print("Memulai pencarian di Sinta... (Proses mungkin memakan waktu)")

    process_names(names, output_file)


if __name__ == "__main__":
    main()