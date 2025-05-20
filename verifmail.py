import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
import time
import random
from openpyxl import load_workbook


def extract_names_from_excel(file_path, sheet_name='Sheet6', name_column='Nama2'):
    """
    Mengekstrak nama dari file author79.xlsx sheet ALL
    """
    try:
        # Menggunakan openpyxl untuk handle file xlsx
        wb = load_workbook(filename=file_path)
        sheet = wb[sheet_name]

        # Cari kolom yang berisi nama
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        col_idx = header_row.index(name_column) if name_column in header_row else 0

        names = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[col_idx]:
                names.append(str(row[col_idx]).strip())

        return list(set(names))  # Hapus duplikat
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []


def search_sinta(name, affiliation="UIN Sunan Kalijaga"):
    """
    Mencari peneliti di Sinta (Science and Technology Index) berdasarkan nama dan afiliasi
    """
    base_url = "https://sinta.kemdikbud.go.id/authors"
    params = {
        "search": name,
        "view": "list"
    }

    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        # Random delay sebelum request
        time.sleep(random.uniform(1, 2))

        response = requests.get(base_url, params=params, headers=headers, timeout=30)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        results = []

        for author in soup.select('.author-list-item'):
            author_name_elem = author.select_one('.author-name a')
            if not author_name_elem:
                continue

            author_name = author_name_elem.text.strip()
            author_url = author_name_elem['href']
            author_affiliation = author.select_one('.affil-name').text.strip() if author.select_one(
                '.affil-name') else ""

            # Cek variasi penulisan UIN Sunan Kalijaga
            if any(term.lower() in author_affiliation.lower() for term in
                   ["UIN Sunan Kalijaga", "UIN Suka", "Sunan Kalijaga", "UIN Yogyakarta"]):
                # Dapatkan ID Sinta dari URL
                sinta_id = re.search(r'/authors/(\d+)', author_url)
                if sinta_id:
                    sinta_id = sinta_id.group(1)
                else:
                    sinta_id = ""

                results.append({
                    'original_name': name,
                    'sinta_name': author_name,
                    'sinta_id': sinta_id,
                    'url': author_url,
                    'affiliation': author_affiliation,
                    'scopus_id': author.get('data-scopus-id', ''),
                    'scholar_id': author.get('data-scholar-id', '')
                })

        return results
    except Exception as e:
        print(f"Error searching Sinta for {name}: {e}")
        return []


def get_sinta_details(sinta_url):
    """
    Mendapatkan detail lengkap dari profil Sinta
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        # Random delay sebelum request
        time.sleep(random.uniform(1, 2))

        response = requests.get(sinta_url, headers=headers, timeout=30)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        # Cari email
        email = ""
        email_element = soup.select_one('.author-email a')
        if email_element:
            email = email_element.text.strip()

        # Cek apakah email menggunakan domain uin-suka.ac.id
        is_uin_email = bool(re.search(r'@(uin-suka\.ac\.id|suka\.ac\.id|uin\.ac\.id)', email, re.IGNORECASE))

        # Dapatkan h-index dan total dokumen
        h_index = ""
        documents = ""
        metrics = soup.select('.author-metrics .metric-value')
        if len(metrics) >= 2:
            h_index = metrics[0].text.strip()
            documents = metrics[1].text.strip()

        return {
            'email': email,
            'is_uin_email': is_uin_email,
            'h_index': h_index,
            'documents': documents
        }
    except Exception as e:
        print(f"Error getting Sinta details for {sinta_url}: {e}")
        return {
            'email': '',
            'is_uin_email': False,
            'h_index': '',
            'documents': ''
        }


def process_names(names, output_file='sinta_results.xlsx'):
    """
    Memproses daftar nama dan menyimpan hasilnya
    """
    results = []
    total_processed = 0
    batch_size = 10  # Jumlah request sebelum istirahat panjang
    batch_count = 0

    for i, name in enumerate(names):
        print(f"Processing {i + 1}/{len(names)}: {name}")
        sinta_accounts = search_sinta(name)

        if not sinta_accounts:
            results.append({
                'Nama Asli': name,
                'Nama di Sinta': '-',
                'ID Sinta': '-',
                'Afiliasi': '-',
                'URL Sinta': '-',
                'Email': '-',
                'Email UIN': '-',
                'H-Index': '-',
                'Total Dokumen': '-',
                'Scopus ID': '-',
                'Scholar ID': '-',
                'Status': 'Tidak ditemukan'
            })
        else:
            for account in sinta_accounts:
                details = get_sinta_details(account['url'])

                results.append({
                    'Nama Asli': account['original_name'],
                    'Nama di Sinta': account['sinta_name'],
                    'ID Sinta': account['sinta_id'],
                    'Afiliasi': account['affiliation'],
                    'URL Sinta': account['url'],
                    'Email': details['email'],
                    'Email UIN': 'Ya' if details['is_uin_email'] else 'Tidak',
                    'H-Index': details['h_index'],
                    'Total Dokumen': details['documents'],
                    'Scopus ID': account['scopus_id'],
                    'Scholar ID': account['scholar_id'],
                    'Status': 'Ditemukan'
                })

        total_processed += 1
        batch_count += 1

        # Istirahat panjang setiap batch_size request
        if batch_count >= batch_size:
            long_delay = random.uniform(1, 30)  # Istirahat 1-3 menit
            print(f"\n--- Menjalankan istirahat panjang selama {long_delay / 60:.1f} menit ---\n")
            time.sleep(long_delay)
            batch_count = 0

    # Simpan hasil ke Excel
    df = pd.DataFrame(results)

    # Urutkan kolom
    columns_order = ['Nama Asli', 'Nama di Sinta', 'ID Sinta', 'Afiliasi', 'URL Sinta',
                     'Email', 'Email UIN', 'H-Index', 'Total Dokumen',
                     'Scopus ID', 'Scholar ID', 'Status']
    df = df[columns_order]

    df.to_excel(output_file, index=False)
    print(f"\nHasil disimpan ke {output_file}")
    print(f"Total diproses: {len(names)} nama")
    print(f"Total ditemukan: {len(df[df['Status'] == 'Ditemukan'])} akun Sinta")


def main():
    # File input dan output
    input_file = "author79.xlsx"
    output_file = "hasil_verifikasi_sinta.xlsx"

    print("Memulai proses ekstraksi nama dari Excel...")
    names = extract_names_from_excel(input_file, sheet_name='Sheet6')

    if not names:
        print("Tidak ditemukan nama di file Excel.")
        return

    print(f"\nDitemukan {len(names)} nama unik untuk diproses.")
    print("Memulai pencarian di Sinta... (Proses mungkin memakan waktu)")

    process_names(names, output_file)


if __name__ == "__main__":
    main()